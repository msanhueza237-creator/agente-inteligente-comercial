from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(".", "").replace(",", "."))
        except ValueError:
            return None
    return None


def positive(value: Any) -> float | None:
    parsed = number(value)
    return parsed if parsed is not None and parsed > 0 else None


def round_value(value: float | None, digits: int = 6) -> float | None:
    return round(value, digits) if value is not None else None


def suggested_catalog(path: Path) -> list[dict[str, Any]]:
    values = load_workbook(path, data_only=True, read_only=True).active
    formulas = load_workbook(path, data_only=False, read_only=True).active
    items: list[dict[str, Any]] = []
    for row_number, (value_row, formula_row) in enumerate(
        zip(
            values.iter_rows(min_row=2, max_col=15, values_only=True),
            formulas.iter_rows(min_row=2, max_col=15, values_only=True),
        ),
        start=2,
    ):
        sku = str(value_row[0] or "").strip()
        name = str(value_row[1] or "").strip()
        if not sku and not name:
            continue
        requested_qty = positive(value_row[2])
        order_qty = positive(value_row[3]) or requested_qty
        unit_fob = positive(value_row[5])
        total_fob = positive(value_row[6])
        cartons = positive(value_row[7])
        gross_weight = positive(value_row[8])
        total_cbm = positive(value_row[9])
        pieces_per_carton = positive(value_row[10])
        carton_weight = positive(value_row[11])
        dimensions = [positive(value_row[index]) for index in (12, 13, 14)]
        if total_cbm is None and cartons and all(dimensions):
            total_cbm = cartons
            for dimension in dimensions:
                total_cbm *= float(dimension)
            total_cbm /= 1_000_000
        if total_fob is None and order_qty and unit_fob:
            total_fob = order_qty * unit_fob
        unit_cbm = total_cbm / order_qty if total_cbm and order_qty else None
        items.append(
            {
                "sku": sku or None,
                "name": name or sku,
                "supplier": "Chinafore",
                "unit": str(value_row[4] or "pcs"),
                "suggested_quantity": round_value(requested_qty, 2),
                "order_quantity": round_value(order_qty, 2),
                "order_multiple": round_value(pieces_per_carton, 2),
                "unit_fob_usd": round_value(unit_fob, 4),
                "total_fob_usd": round_value(total_fob, 2),
                "cartons": round_value(cartons, 2),
                "gross_weight_kg": round_value(gross_weight, 2),
                "total_cbm": round_value(total_cbm),
                "unit_cbm": round_value(unit_cbm, 8),
                "carton_weight_kg": round_value(carton_weight, 2),
                "carton_dimensions_cm": [round_value(value, 2) for value in dimensions]
                if all(dimensions)
                else None,
                "source_document": path.name,
                "source_kind": "purchase_suggestion",
                "source_row": row_number,
                "source_drive_folder": "agente comercio exterior/chinafore proveedor",
                "source_verified_at": "2026-07-31",
                "volume_evidence": "document" if positive(value_row[9]) else "carton_calculation" if total_cbm else "missing",
                "formula_reference": str(formula_row[9] or "") or None,
            }
        )
    return items


def current_order_catalog(path: Path) -> list[dict[str, Any]]:
    values = load_workbook(path, data_only=True, read_only=True).active
    items: list[dict[str, Any]] = []
    for row_number, row in enumerate(
        values.iter_rows(min_row=2, max_col=14, values_only=True), start=2
    ):
        name = str(row[1] or row[0] or "").strip()
        if not name:
            continue
        quantity = positive(row[2])
        unit_fob = positive(row[4])
        total_fob = positive(row[5])
        cartons = positive(row[6])
        gross_weight = positive(row[7])
        total_cbm = positive(row[8])
        pieces_per_carton = positive(row[9])
        carton_weight = positive(row[10])
        dimensions = [positive(row[index]) for index in (11, 12, 13)]
        if total_cbm is None and cartons and all(dimensions):
            total_cbm = cartons
            for dimension in dimensions:
                total_cbm *= float(dimension)
            total_cbm /= 1_000_000
        if total_fob is None and quantity and unit_fob:
            total_fob = quantity * unit_fob
        unit_cbm = total_cbm / quantity if total_cbm and quantity else None
        items.append(
            {
                "sku": str(row[0] or "").strip() or None,
                "name": name,
                "supplier": "Chinafore",
                "unit": str(row[3] or "pcs"),
                "suggested_quantity": round_value(quantity, 2),
                "order_quantity": round_value(quantity, 2),
                "order_multiple": round_value(pieces_per_carton, 2),
                "unit_fob_usd": round_value(unit_fob, 4),
                "total_fob_usd": round_value(total_fob, 2),
                "cartons": round_value(cartons, 2),
                "gross_weight_kg": round_value(gross_weight, 2),
                "total_cbm": round_value(total_cbm),
                "unit_cbm": round_value(unit_cbm, 8),
                "carton_weight_kg": round_value(carton_weight, 2),
                "carton_dimensions_cm": [round_value(value, 2) for value in dimensions]
                if all(dimensions)
                else None,
                "source_document": path.name,
                "source_kind": "current_order",
                "source_row": row_number,
                "source_drive_folder": "agente comercio exterior/chinafore proveedor",
                "source_verified_at": "2026-07-31",
                "volume_evidence": "document" if positive(row[8]) else "carton_calculation" if total_cbm else "missing",
            }
        )
    return items


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--suggestion", type=Path, required=True)
    parser.add_argument("--current-order", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    items = [*suggested_catalog(args.suggestion), *current_order_catalog(args.current_order)]
    payload = {
        "schema_version": 1,
        "generated_at": date.today().isoformat(),
        "supplier": "Chinafore",
        "currency": "USD",
        "items": items,
        "coverage": {
            "products": len(items),
            "with_sku": sum(bool(item["sku"]) for item in items),
            "with_fob": sum(bool(item["unit_fob_usd"]) for item in items),
            "with_cbm": sum(bool(item["unit_cbm"]) for item in items),
        },
        "sources": [
            {"file": args.suggestion.name, "kind": "purchase_suggestion"},
            {"file": args.current_order.name, "kind": "current_order"},
        ],
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
